import os
import glob
import openpyxl
import json

for FileList in glob.glob('*.xlsx'):
    path = os.path.abspath(os.getcwd())
    base_new = path
    path = path + '/' + FileList
    exel_file = openpyxl.load_workbook(path).active
    subject = exel_file.cell(row=2, column=1).value
    level = exel_file.cell(row=2, column=2).value
    final_dict = {}
    subject_dict = {}
    curr_dict = {}

    for i in range(2, exel_file.max_row + 1):
        if exel_file.cell(row=i, column=1).value != None and exel_file.cell(row=i, column=1).value != subject:
            subject_dict[level] = curr_dict
            final_dict[subject] = subject_dict
            subject = exel_file.cell(row=i, column=1).value
            level = exel_file.cell(row=i, column=2).value
            subject_dict = {}
            curr_dict = {}
        if exel_file.cell(row=i, column=2).value != None and exel_file.cell(row=i, column=2).value != level:
            subject_dict[level] = curr_dict
            level = exel_file.cell(row=i, column=2).value
            curr_dict = {}
        curr_dict[exel_file.cell(row=i, column=3).value] = {"Variable1Min": str(exel_file.cell(row=i, column=6).value),
                                                            "Variable1Max": str(exel_file.cell(row=i, column=7).value),
                                                            "Variable2Min": str(exel_file.cell(row=i, column=8).value),
                                                            "Variable2Max": str(exel_file.cell(row=i, column=9).value),
                                                            "Correct": str(exel_file.cell(row=i, column=10).value),
                                                            "Incorrect1": str(exel_file.cell(row=i, column=11).value),
                                                            "Incorrect2": str(exel_file.cell(row=i, column=12).value),
                                                            "Incorrect3": str(exel_file.cell(row=i, column=13).value)}
                                                            #**({"Image": exel_file.cell(row=i, column=8).value} if exel_file.cell(row=i, column=8).value is not None else {"Image": ""}),
                                                            #**({"Suggestion": exel_file.cell(row=i, column=9).value} if exel_file.cell(row=i,column=9).value is not None else {"Suggestion": ""})}

    subject_dict[level] = curr_dict
    final_dict[subject] = subject_dict
    file_name = FileList.split('.', 1)[0]
    file_path = file_name = base_new + "\\" + file_name + ".json"
    with open(file_path, 'w') as resultFile:
        json.dump(final_dict, resultFile, indent=4)
        resultFile.write('\n')