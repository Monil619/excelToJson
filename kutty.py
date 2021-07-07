import os
import glob
import openpyxl
import json
for FileList in glob.glob('*.xlsx'):
    path = os.path.abspath(os.getcwd())
    base_new = path
    path = path + '/' + FileList
    exel_file = openpyxl.load_workbook(path).active
    subject = exel_file.cell(row=2, column=1).value
    level = exel_file.cell(row=2, column=2).value
    final_dict = {}
    subject_dict = {}
    curr_dict = {}
    for i in range(2, exel_file.max_row + 1):
        if exel_file.cell(row=i, column=1).value != None and exel_file.cell(row=i, column=1).value != subject:
            subject_dict[level] = curr_dict
            final_dict[subject] = subject_dict
            subject = exel_file.cell(row=i, column=1).value
            level = exel_file.cell(row=i, column=2).value
            subject_dict = {}
            curr_dict = {}
        if exel_file.cell(row=i, column=2).value != None and exel_file.cell(row=i, column=2).value != level:
            subject_dict[level] = curr_dict
            level = exel_file.cell(row=i, column=2).value
            curr_dict = {}

        q_part1_front = str(exel_file.cell(row=i, column=3).value).lstrip()
        q_part1_back = q_part1_front.rstrip()
        q_part2_front = str(exel_file.cell(row=i, column=4).value).lstrip()
        q_part2_back = q_part2_front.rstrip()
        q_part3_front = str(exel_file.cell(row=i, column=5).value).lstrip()
        q_part3_back = q_part3_front.rstrip()

        question =  q_part1_back + " " + q_part2_back + " " + q_part3_back

        curr_dict[question] = {"Variable1Min": str(exel_file.cell(row=i, column=6).value) if not str(exel_file.cell(row=i, column=6).value).isspace() else "",
                                                            "Variable1Max": str(exel_file.cell(row=i, column=7).value) if not str(exel_file.cell(row=i, column=7).value).isspace() else "",
                                                             "Variable2Min": str(exel_file.cell(row=i, column=8).value) if not str(exel_file.cell(row=i, column=8).value).isspace() else "",
                                                            "Variable2Max": str(exel_file.cell(row=i, column=9).value) if not str(exel_file.cell(row=i, column=9).value).isspace() else "",
                                                            "Correct": str(exel_file.cell(row=i, column=10).value) if not str(exel_file.cell(row=i, column=10).value).isspace() else "",
                                                            "Incorrect1": str(exel_file.cell(row=i, column=11).value) if not str(exel_file.cell(row=i,column=11).value).isspace() else "",
                                                            "Incorrect2": str(exel_file.cell(row=i, column=12).value) if not str(exel_file.cell(row=i,column=12).value).isspace() else "",
                                                            "Incorrect3": str(exel_file.cell(row=i,column=13).value) if not str(exel_file.cell(row=i,column=13).value).isspace() else ""}

    subject_dict[level] = curr_dict
    final_dict[subject] = subject_dict
    file_name = FileList.split('.', 1)[0]
    file_path = file_name = base_new + "\\" + file_name + ".json"
    with open(file_path, 'w') as resultFile:
        json.dump(final_dict, resultFile, indent=4)
        resultFile.write('\n')